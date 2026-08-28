#!/usr/bin/env python3
"""Merge hand-collected readings from an .xlsx sheet into the CSV history.

Before the collector existed the counters were copied off the page by hand into
a spreadsheet. This folds those readings into docs/data/ so the history reaches
further back than the automation does.

    python tools/import_manual.py ~/Downloads/ESO.xlsx            # preview only
    python tools/import_manual.py ~/Downloads/ESO.xlsx --write    # actually merge

The sheet is read by its labels, not by cell positions: every cell that looks
like "Klientų pranešimų skaičius: 5067" is matched against the four known
counter names, and the timestamp comes from the datetime at the top of that
cell's column. Rows and columns can therefore move without breaking the import.

Two rules the import never bends:

* Timestamps in the sheet are Vilnius wall-clock. They are converted to UTC,
  because that is what the CSV stores.
* A counter the sheet does not carry is written empty, never zero. An invented
  zero is indistinguishable from a measured one once it is in the file.
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "collector"))

from collect import FIELDS, rebuild_index  # noqa: E402

DATA_DIR = ROOT / "docs" / "data"


def _zoneinfo():
    """The system zone database when it exists - it is the authority."""
    try:
        from zoneinfo import ZoneInfo

        return ZoneInfo("Europe/Vilnius")
    except Exception:
        # Windows ships no tz database and `tzdata` is not installed by default.
        # Rather than add a dependency to a repo that is otherwise stdlib-only,
        # fall back to the EU rule below.
        return None


TZ = _zoneinfo()


def _last_sunday(year: int, month: int) -> date:
    following = date(year + 1, 1, 1) if month == 12 else date(year, month + 1, 1)
    last = following - timedelta(days=1)
    return last - timedelta(days=(last.weekday() - 6) % 7)


def _in_summer_time(moment_utc: datetime) -> bool:
    """EU rule, unchanged since 2002: last Sunday of March 01:00 UTC to last
    Sunday of October 01:00 UTC. Lithuania has followed it throughout."""
    year = moment_utc.year
    start = datetime.combine(_last_sunday(year, 3), datetime.min.time()) + timedelta(hours=1)
    end = datetime.combine(_last_sunday(year, 10), datetime.min.time()) + timedelta(hours=1)
    return start <= moment_utc.replace(tzinfo=None) < end


def local_to_utc(naive_local: datetime) -> datetime:
    """Vilnius wall-clock to UTC."""
    if TZ is not None:
        return naive_local.replace(tzinfo=TZ).astimezone(timezone.utc)

    for offset in (3, 2):
        candidate = naive_local - timedelta(hours=offset)
        if _in_summer_time(candidate) == (offset == 3):
            return candidate.replace(tzinfo=timezone.utc)
    # The hour that happens twice each autumn: read it as standard time.
    return (naive_local - timedelta(hours=2)).replace(tzinfo=timezone.utc)

# Matched by prefix, so the parenthetical note ESO appends to `n` does not matter.
LABELS = {
    "klientų pranešimų skaičius": "k",
    "sutrikimų skaičius": "c",
    "dėl sutrikimų atjungtų klientų skaičius": "n",
    "dėl planinių darbų atjungtų klientų skaičius": "p",
}


def metric_for(label: str) -> str | None:
    cleaned = label.strip().lower()
    for prefix, key in LABELS.items():
        if cleaned.startswith(prefix):
            return key
    return None


def read_sheet(path: Path) -> list[dict]:
    workbook = openpyxl.load_workbook(path, data_only=True)
    samples: dict[datetime, dict] = {}

    for sheet in workbook.worksheets:
        # A column belongs to the moment named by the topmost datetime in it.
        column_time: dict[int, datetime] = {}
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, datetime) and cell.column not in column_time:
                    column_time[cell.column] = cell.value

        for row in sheet.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str) or ":" not in cell.value:
                    continue
                label, _, tail = cell.value.rpartition(":")
                key = metric_for(label)
                if key is None:
                    continue
                digits = re.sub(r"[^\d-]", "", tail)
                if not digits:
                    continue
                moment = column_time.get(cell.column)
                if moment is None:
                    continue
                samples.setdefault(moment, {})[key] = int(digits)

    rows = []
    for moment, values in sorted(samples.items()):
        utc = local_to_utc(moment)
        row = {"ts_utc": utc.strftime("%Y-%m-%dT%H:%M:%SZ"), "server_ts": ""}
        for key in ("k", "c", "n", "p"):
            row[key] = values.get(key, "")
        rows.append(row)
    return rows


def read_existing(path: Path) -> list[dict]:
    if not path.exists():
        return []
    with path.open(encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))


def write_rows(path: Path, rows: list[dict]) -> None:
    with path.open("w", encoding="utf-8", newline="\n") as handle:
        writer = csv.DictWriter(handle, fieldnames=FIELDS, lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--write", action="store_true", help="merge for real")
    parser.add_argument("--data-dir", type=Path, default=DATA_DIR)
    args = parser.parse_args()

    imported = read_sheet(args.workbook)
    if not imported:
        print("в книге не нашлось ни одного распознанного показателя")
        return

    print(f"в книге: {len(imported)} замеров, "
          f"{imported[0]['ts_utc']} … {imported[-1]['ts_utc']}\n")
    for row in imported:
        missing = [key for key in ("k", "c", "n", "p") if row[key] == ""]
        note = f"  (нет: {', '.join(missing)})" if missing else ""
        print(f"  {row['ts_utc']}  k={row['k'] or '-':>6} c={row['c'] or '-':>5} "
              f"n={row['n'] or '-':>7} p={row['p'] or '-':>5}{note}")

    by_month: dict[str, list[dict]] = {}
    for row in imported:
        by_month.setdefault(row["ts_utc"][:7], []).append(row)

    total_added = 0
    for month, new_rows in sorted(by_month.items()):
        path = args.data_dir / f"{month}.csv"
        existing = read_existing(path)
        seen = {row["ts_utc"] for row in existing}
        fresh = [row for row in new_rows if row["ts_utc"] not in seen]
        skipped = len(new_rows) - len(fresh)
        total_added += len(fresh)

        print(f"\n{path.name}: было {len(existing)}, добавится {len(fresh)}"
              + (f", пропущено как дубликаты {skipped}" if skipped else ""))

        if not args.write or not fresh:
            continue

        merged = sorted(existing + fresh, key=lambda row: row["ts_utc"])
        write_rows(path, merged)
        print(f"  записано, теперь строк: {len(merged)}")

    if not args.write:
        print("\nпредпросмотр: ничего не записано, добавьте --write")
        return

    if total_added:
        latest = read_existing(sorted(args.data_dir.glob("[0-9]*.csv"))[-1])[-1]
        rebuild_index(args.data_dir, latest)
        print("\nindex.json и latest.json пересобраны")


if __name__ == "__main__":
    main()
